"""從 Excel 讀取 AT 測試步驟與設定。"""

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import load_workbook

from custom_expects import resolve_custom_expect

STEPS_SHEET = "Steps"
CONFIG_SHEET = "Config"
REFERENCE_SHEET = "Reference"

REQUIRED_STEP_COLUMNS = [
    "test_id",
    "step",
    "enabled",
    "cmd",
    "expect_type",
    "expect_value",
    "timeout",
]

OPTIONAL_STEP_COLUMNS = [
    "expect_label",
    "idle_timeout",
    "wait_after",
    "reconnect_after",
    "run_on_failure",
    "check_ttff",
    "note",
]

BOOL_TRUE = {"v", "y", "yes", "true", "1", "是"}
BOOL_FALSE = {"n", "no", "false", "0", "否", ""}


class ExcelLoadError(Exception):
    """Excel 格式或內容錯誤。"""


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_bool(value: Any, field_name: str, row_no: int) -> bool:
    text = _cell_str(value).lower()
    if text in BOOL_TRUE:
        return True
    if text in BOOL_FALSE:
        return False
    raise ExcelLoadError(
        f"Steps 第 {row_no} 列：{field_name} 必須為 V/Y（是）或 N/空白（否）（目前: {value!r}）"
    )


def _parse_float(value: Any, field_name: str, row_no: int, required: bool = True) -> Optional[float]:
    text = _cell_str(value)
    if not text:
        if required:
            raise ExcelLoadError(f"Steps 第 {row_no} 列：{field_name} 不可為空")
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise ExcelLoadError(
            f"Steps 第 {row_no} 列：{field_name} 必須為數字（目前: {value!r}）"
        ) from exc


def _parse_idle_timeout(value: Any, row_no: int) -> Optional[float]:
    text = _cell_str(value).lower()
    if not text:
        return 0.3
    if text in {"none", "null", "-", "無"}:
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise ExcelLoadError(
            f"Steps 第 {row_no} 列：idle_timeout 必須為數字或 none（目前: {value!r}）"
        ) from exc


def _read_sheet_rows(path: Union[str, Path], sheet_name: str) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ExcelLoadError(f"找不到工作表 {sheet_name!r}（現有: {workbook.sheetnames}）")

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return []

    headers = [_cell_str(h) for h in rows[0]]
    if not any(headers):
        return []

    result: List[Dict[str, Any]] = []
    for row_no, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or _cell_str(cell) == "" for cell in row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[idx] if idx < len(row) else None
        item["_row_no"] = row_no
        result.append(item)
    return result


def load_config_from_excel(
    path: Union[str, Path],
    sheet: str = CONFIG_SHEET,
) -> Dict[str, str]:
    """讀取 Config 工作表（key / value 兩欄）。"""
    rows = _read_sheet_rows(path, sheet)
    config: Dict[str, str] = {}
    for row in rows:
        key = _cell_str(row.get("key"))
        if not key or key.startswith("#"):
            continue
        config[key] = _cell_str(row.get("value"))
    return config


def list_test_ids(
    path: Union[str, Path],
    sheet: str = STEPS_SHEET,
    enabled_only: bool = True,
) -> List[str]:
    """列出 Excel 中所有 test_id（依出現順序、去重）。"""
    rows = _read_sheet_rows(path, sheet)
    seen = set()
    test_ids: List[str] = []
    for row in rows:
        test_id = _cell_str(row.get("test_id"))
        if not test_id:
            continue
        if enabled_only and not _parse_bool(row.get("enabled"), "enabled", row["_row_no"]):
            continue
        if test_id not in seen:
            seen.add(test_id)
            test_ids.append(test_id)
    return test_ids


def _build_expect(step_row: Dict[str, Any]) -> tuple:
    row_no = step_row["_row_no"]
    expect_type = _cell_str(step_row.get("expect_type")).lower()
    expect_value = _cell_str(step_row.get("expect_value"))
    expect_label = _cell_str(step_row.get("expect_label"))

    if expect_type == "contains":
        if not expect_value:
            raise ExcelLoadError(f"Steps 第 {row_no} 列：contains 類型需填 expect_value")
        label = expect_label or expect_value
        return expect_value, label

    if expect_type == "regex":
        if not expect_value:
            raise ExcelLoadError(f"Steps 第 {row_no} 列：regex 類型需填 expect_value")
        pattern = re.compile(expect_value)
        label = expect_label or expect_value
        return pattern, label

    if expect_type == "custom":
        checker, default_label = resolve_custom_expect(expect_value)
        label = expect_label or default_label
        return checker, label

    raise ExcelLoadError(
        f"Steps 第 {row_no} 列：未知 expect_type {expect_type!r}"
        f"（可用: contains, regex, custom）"
    )


def _row_to_step_dict(row: Dict[str, Any]) -> dict:
    row_no = row["_row_no"]

    for col in REQUIRED_STEP_COLUMNS:
        if col not in row:
            raise ExcelLoadError(f"Steps 缺少必要欄位: {col}")
        if col != "enabled" and col != "expect_value" and not _cell_str(row.get(col)):
            if col == "step":
                raise ExcelLoadError(f"Steps 第 {row_no} 列：step 不可為空")
            if col in ("cmd", "expect_type", "timeout"):
                raise ExcelLoadError(f"Steps 第 {row_no} 列：{col} 不可為空")

    expect, expect_label = _build_expect(row)
    timeout = _parse_float(row.get("timeout"), "timeout", row_no)
    assert timeout is not None

    step: dict = {
        "cmd": _cell_str(row.get("cmd")),
        "expect": expect,
        "expect_label": expect_label,
        "timeout": timeout,
        "idle_timeout": _parse_idle_timeout(row.get("idle_timeout"), row_no),
    }

    wait_after = _parse_float(row.get("wait_after"), "wait_after", row_no, required=False)
    if wait_after is not None:
        step["wait_after"] = wait_after

    if _parse_bool(row.get("reconnect_after"), "reconnect_after", row_no):
        step["reconnect_after"] = True

    if _parse_bool(row.get("run_on_failure"), "run_on_failure", row_no):
        step["run_on_failure"] = True

    if _parse_bool(row.get("check_ttff"), "check_ttff", row_no):
        step["check_ttff"] = True

    note = _cell_str(row.get("note"))
    if note:
        step["note"] = note

    return step


def load_steps_from_excel(
    path: Union[str, Path],
    test_id: str,
    sheet: str = STEPS_SHEET,
    enabled_only: bool = True,
) -> List[dict]:
    """
    從 Excel 載入指定 test_id 的 AT 步驟清單。

    回傳格式與 GNSS 腳本 build_at_steps() 相同，可直接傳入 run_at_sequence()。
    """
    path = Path(path)
    if not path.is_file():
        raise ExcelLoadError(f"找不到 Excel 檔案: {path}")

    rows = _read_sheet_rows(path, sheet)
    if not rows:
        raise ExcelLoadError(f"工作表 {sheet!r} 沒有資料列")

    matched: List[Dict[str, Any]] = []
    for row in rows:
        row_test_id = _cell_str(row.get("test_id"))
        if row_test_id != test_id:
            continue
        if enabled_only and not _parse_bool(row.get("enabled"), "enabled", row["_row_no"]):
            continue
        matched.append(row)

    if not matched:
        available = list_test_ids(path, sheet=sheet, enabled_only=False)
        raise ExcelLoadError(
            f"找不到 test_id={test_id!r} 的啟用步驟（Excel 內 test_id: {available}）"
        )

    matched.sort(key=lambda r: int(float(_cell_str(r.get("step")) or "0")))
    return [_row_to_step_dict(row) for row in matched]


def validate_excel(path: Union[str, Path], sheet: str = STEPS_SHEET) -> List[str]:
    """驗證 Excel 格式（含 enabled=N 的步驟），回傳所有 test_id；有錯誤則拋 ExcelLoadError。"""
    test_ids = list_test_ids(path, sheet=sheet, enabled_only=False)
    for test_id in test_ids:
        # 驗證用 enabled_only=False，避免「全部停用」的測項被誤判為格式錯誤
        load_steps_from_excel(path, test_id=test_id, sheet=sheet, enabled_only=False)
    return test_ids
