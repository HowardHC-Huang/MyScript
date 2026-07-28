"""產生 AT_Testcases_Template.xlsx 範本檔。"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent / "testcases" / "AT_Testcases_Template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header_row(sheet, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    sheet.freeze_panes = "A2"


def _autosize(sheet, col_count: int) -> None:
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in sheet[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)


def _build_config_sheet(wb: Workbook) -> None:
    sheet = wb.create_sheet("Config", 0)
    headers = ["key", "value", "說明"]
    sheet.append(headers)
    rows = [
        ["port", "COM14", "Serial port（執行時可用命令列覆寫）"],
        ["baudrate", "115200", "鮑率"],
        ["serial_timeout", "1.0", "pyserial read timeout 秒數"],
        ["rounds", "1", "測試回合數"],
        ["default_step_wait", "0", "步驟間預設等待秒數（各步驟 wait_after 可覆寫）"],
        ["reconnect_max_wait", "60", "reconnect_after=V 時，重連最多等待秒數"],
        ["test_id", "", "單一 test_id；留空且 test_ids 也留空時，執行全部 enabled=V 測項"],
        [
            "test_ids",
            "",
            "多個 test_id（逗號分隔）；有值時優先於 test_id",
        ],
    ]
    for row in rows:
        sheet.append(row)
    _style_header_row(sheet, len(headers))
    _autosize(sheet, len(headers))


def _build_steps_sheet(wb: Workbook) -> None:
    sheet = wb.create_sheet("Steps", 1)
    headers = [
        "test_id",
        "step",
        "enabled",
        "cmd",
        "expect_type",
        "expect_value",
        "expect_label",
        "timeout",
        "idle_timeout",
        "wait_after",
        "reconnect_after",
        "run_on_failure",
        "check_ttff",
        "note",
    ]
    sheet.append(headers)

    # 與 GNSS_XTRA_ColdStart_STRESS_0001.py build_at_steps() 對應的範例
    rows = [
        [
            "TC_GNSS_XTRA_ColdStart",
            1,
            "V",
            'AT!ENTERCND="A710"',
            "contains",
            "OK",
            "",
            5,
            "",
            0,
            "N",
            "N",
            "N",
            "進入 CND 模式",
        ],
        [
            "TC_GNSS_XTRA_ColdStart",
            2,
            "V",
            "AT!GPSCLRASSIST=1,1,1,1,1",
            "contains",
            "OK",
            "",
            10,
            "",
            0,
            "N",
            "N",
            "N",
            "清除 assist data",
        ],
        [
            "TC_GNSS_XTRA_ColdStart",
            3,
            "V",
            "AT!GPSXTRASTATUS?",
            "contains",
            "1980",
            "",
            30,
            "none",
            0,
            "N",
            "N",
            "N",
            "XTRA 應為初始狀態",
        ],
        [
            "TC_GNSS_XTRA_ColdStart",
            4,
            "V",
            "AT!GPSXTRAINITDNLD",
            "contains",
            "Xtra command sent successfully",
            "",
            60,
            "none",
            0,
            "N",
            "N",
            "N",
            "觸發 XTRA 下載",
        ],
        [
            "TC_GNSS_XTRA_ColdStart",
            5,
            "V",
            "AT!GPSXTRASTATUS?",
            "custom",
            "current_year_month",
            "",
            30,
            "none",
            0,
            "N",
            "N",
            "N",
            "XTRA 日期應為今年今月",
        ],
        [
            "TC_GNSS_XTRA_ColdStart",
            6,
            "V",
            "AT!GPSFIX=1,255,255",
            "contains",
            "OK",
            "",
            30,
            "",
            5,
            "N",
            "N",
            "N",
            "Cold start fix",
        ],
        [
            "TC_GNSS_XTRA_ColdStart",
            7,
            "V",
            "AT!GPSSTATUS?",
            "custom",
            "gps_status_ttff",
            "",
            30,
            "none",
            0,
            "N",
            "N",
            "V",
            "檢查 TTFF",
        ],
        [
            "TC_GNSS_XTRA_ColdStart",
            8,
            "V",
            "AT!GPSEND=0",
            "contains",
            "OK",
            "",
            10,
            "",
            0,
            "N",
            "V",
            "N",
            "收尾；失敗時仍執行",
        ],
        [
            "TC_Smoke_AT",
            1,
            "V",
            "AT",
            "contains",
            "OK",
            "",
            2,
            "",
            0,
            "N",
            "N",
            "N",
            "基本握手",
        ],
        [
            "TC_Smoke_AT",
            2,
            "V",
            "ATI",
            "regex",
            "OK",
            "",
            2,
            "",
            0,
            "N",
            "N",
            "N",
            "查詢版本",
        ],
    ]
    for row in rows:
        sheet.append(row)
    _style_header_row(sheet, len(headers))
    _autosize(sheet, len(headers))


def _build_reference_sheet(wb: Workbook) -> None:
    sheet = wb.create_sheet("Reference", 2)
    headers = ["欄位", "說明", "範例"]
    sheet.append(headers)
    rows = [
        ["test_id", "測試案例 ID，同 ID 多列組成一個序列", "TC_GNSS_XTRA_ColdStart"],
        ["step", "步驟順序（整數）", "1"],
        ["enabled", "V=執行 / N=略過", "V"],
        ["cmd", "AT 指令（不含 \\r\\n）", "AT"],
        ["expect_type", "contains | regex | custom", "contains"],
        ["expect_value", "預期內容或 custom 名稱", "OK / current_year_month"],
        ["expect_label", "報告顯示用（可空）", "TTFF (sec) < 5"],
        ["timeout", "此步驟最長等待秒數", "30"],
        ["idle_timeout", "空=0.3；none=等到 timeout", "none"],
        ["wait_after", "本步驟完成後等待秒數", "5"],
        ["reconnect_after", "V=等待後關閉並重連 serial（模組會重啟時用）", "V"],
        ["run_on_failure", "V=前步失敗時仍執行此步", "V"],
        ["check_ttff", "V=額外寫 TTFF 判定 log", "V"],
        ["note", "備註", ""],
        ["", "", ""],
        ["custom 名稱", "說明", ""],
        ["current_year_month", "回應含今年+空格+兩位數月份", ""],
        ["gps_status_ttff", "OK 且 TTFF < 5", ""],
        ["gps_status_ttff:3", "OK 且 TTFF < 3", ""],
    ]
    for row in rows:
        sheet.append(row)
    _style_header_row(sheet, len(headers))
    _autosize(sheet, len(headers))


def create_template(output: Path = OUTPUT) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    _build_config_sheet(wb)
    _build_steps_sheet(wb)
    _build_reference_sheet(wb)
    wb.save(output)
    return output


def main() -> None:
    path = create_template()
    print(f"已建立範本: {path}")


if __name__ == "__main__":
    main()
